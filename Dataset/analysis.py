import re
import csv

import openpyxl


def count_entries(data_str):
    entry_count = len(re.findall(r'\{.*?\}', data_str))

    # 计算entry的数量并返回
    return entry_count


def count_entry_occurrences(file_path):
    # 记录句子长度
    text_len_occurrences = {}
    # 用于记录 entry 出现次数的字典
    aspect_occurrences = {}

    # 打开CSV文件
    with open(file_path, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)

        # 遍历每一行数据
        for row in csv_reader:
            # 获取raw_text的长度
            raw_text = row['raw_text']
            raw_text_len = len(raw_text)
            if raw_text_len in text_len_occurrences:
                text_len_occurrences[raw_text_len] += 1
            else:
                text_len_occurrences[raw_text_len] = 1

            # 获取aspect列的JSON数据
            aspect_json = row['aspectTerms']

            # 计算entry的数量
            entry_count = count_entries(aspect_json)

            # 统计 entry 出现次数
            if entry_count in aspect_occurrences:
                aspect_occurrences[entry_count] += 1
            else:
                aspect_occurrences[entry_count] = 1

    return text_len_occurrences, aspect_occurrences


# print("每个 entry 出现的次数统计:")
raw_text_occurrences, aspect_occurrences = count_entry_occurrences("./SemEval14/Train/Laptops_Train.csv")
# raw_text_occurrences = sorted(raw_text_occurrences.keys())
# aspect_occurrences = sorted(aspect_occurrences.keys())
# for count, occurrences in raw_text_occurrences.items():
#     print(f"raw text长度为 {count},出现: {occurrences} 次")
# for count, occurrences in aspect_occurrences.items():
#     print(f"AspectTerm有 {count} 项出现: {occurrences} 次")

# file_path = './SemEval14/Train/Laptops_Train.csv'
# raw_text_occurrences, aspect_occurrences = count_entry_occurrences(file_path)
raw_text_list = sorted(raw_text_occurrences.keys())
aspect_list = sorted(aspect_occurrences.keys())

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()

# 写入第一个字典到第一个 sheet
sheet1 = workbook.active
sheet1.title = "raw text"
sheet1.append(['raw text length', 'times'])
for key in raw_text_list:
    value = raw_text_occurrences.get(key)
    sheet1.append([key, value])
    print(f"raw text长度为 {key},出现:{value} 次")

# 创建一个新的 sheet 并写入第二个字典
sheet2 = workbook.create_sheet(title="aspect")
sheet2.append(['number of aspect term', 'times'])
for key in aspect_list:
    value = aspect_occurrences.get(key)
    sheet2.append([key, value])
    print(f"aspect term个数为 {key},出现: {aspect_occurrences.get(key)} 次")

# 保存工作簿
xlsx_file = "analysis_laptops.xlsx"
workbook.save(xlsx_file)

print("XLSX 文件已生成:", xlsx_file)

raw_text_occurrences, aspect_occurrences = count_entry_occurrences("./SemEval14/Train/Restaurants_Train.csv")
# raw_text_occurrences = sorted(raw_text_occurrences.keys())
# aspect_occurrences = sorted(aspect_occurrences.keys())
# for count, occurrences in raw_text_occurrences.items():
#     print(f"raw text长度为 {count},出现: {occurrences} 次")
# for count, occurrences in aspect_occurrences.items():
#     print(f"AspectTerm有 {count} 项出现: {occurrences} 次")

# file_path = './SemEval14/Train/Laptops_Train.csv'
# raw_text_occurrences, aspect_occurrences = count_entry_occurrences(file_path)
raw_text_list = sorted(raw_text_occurrences.keys())
aspect_list = sorted(aspect_occurrences.keys())

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()

# 写入第一个字典到第一个 sheet
sheet1 = workbook.active
sheet1.title = "raw text"
sheet1.append(['raw text length', 'times'])
for key in raw_text_list:
    value = raw_text_occurrences.get(key)
    sheet1.append([key, value])
    print(f"raw text长度为 {key},出现:{value} 次")

# 创建一个新的 sheet 并写入第二个字典
sheet2 = workbook.create_sheet(title="aspect")
sheet2.append(['number of aspect term', 'times'])
for key in aspect_list:
    value = aspect_occurrences.get(key)
    sheet2.append([key, value])
    print(f"aspect term个数为 {key},出现: {aspect_occurrences.get(key)} 次")

# 保存工作簿
xlsx_file = "analysis_restaurants.xlsx"
workbook.save(xlsx_file)

print("XLSX 文件已生成:", xlsx_file)
